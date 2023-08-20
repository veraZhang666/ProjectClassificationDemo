import openpyxl
import config

def labelZH2digit(detect_result):
    '''
    detect_result : list of list
    return  y_true,y_pred  y_true:[0,4,5,6,1...], y_pred [1,2,3,0...]
    '''
    y_true = []
    y_pred = []

    for record in detect_result:
        y_true.append( config.EN2NUM[config.ZH2EN[record[-2]]])
        y_pred.append(config.EN2NUM[config.ZH2EN[record[-1]]])

    assert len(y_true) == len(y_pred)

    return y_true,y_pred




def check_test_data_format(test_data):
    '''
    test_Data; list of tuples []
    '''
    for txt,label in test_data:
        if len(txt)<200:
            print('字数必须大于200')
            return False
        if label not in config.LABELS:
            print('只允许填写在限制范围内的标签')
            return False

    return True


def read_xlsx_excel(xlsx_path):
    '''
    参数：
        url:xlsx 文件路径
    返回：
        list_of_samples: [（‘txt’,体育），（'txt','科技'）...]
    '''
    # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
    workbook = openpyxl.load_workbook(xlsx_path)
    # 只读取第一个sheet的数据
    name = workbook.sheetnames[0]
    sheet = workbook[name]
    # 定义列表存储表格数据
    list_of_samples = []
    # 遍历表格的每一行

    # 遍历除了第一行的所有行
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row = row[1:3]
        txt,label = row[0].value,row[1].value
        text = txt.strip()
        label = label.strip()
        list_of_samples.append((text,label))
    print(list_of_samples)

    res = check_test_data_format(list_of_samples)
    if res:
        return list_of_samples
    else:
        print('检查xls文件格式是否错误')
        return None




if __name__ == '__main__':
    path = './uploads/predictData.xlsx'
    # print(read_xlsx_excel(path))
    read_xlsx_excel(path)


